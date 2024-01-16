import functools
from decimal import Decimal
from msilib.schema import Font
from django.db import transaction
from django.db.models import Q
from django.db.models import Max
from openpyxl.workbook import Workbook
from rest_framework import viewsets, filters
from django.http import Http404
from .paginations import *
from rest_framework.decorators import action
from .serializers import *
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import CustomTokenObtainPairSerializer
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum
from django.utils import timezone
from django.db.models import F, IntegerField
from django.db.models.functions import Cast
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from openpyxl.styles import Border, Side
from openpyxl.styles import Font


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer




# Отчет с даты по дату + поставщик в ексель документ
# 127.0.0.1:8000/api/report/?start_date=02.01.2024&end_date=03.01.2024&provider_id=2

class ReportViewSet(viewsets.ViewSet):

    def list(self, request):
        # Получаем значения параметров 'start_date', 'end_date' и 'provider_id' из запроса
        start_date_str = request.query_params.get('start_date', None)
        end_date_str = request.query_params.get('end_date', None)
        provider_id = request.query_params.get('provider_id', None)

        # Если не указаны даты, возвращаем ошибку
        if not start_date_str or not end_date_str:
            return Response({'error': 'Необходимо указать start_date и end_date.'}, status=400)

        try:
            # Преобразуем значения параметров в даты
            start_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            # В случае некорректного формата даты возвращаем ошибку
            return Response({'error': 'Неверный формат даты. Пожалуйста, используйте формат DD.MM.YYYY.'}, status=400)

        # Если указан поставщик, вычисляем сальдо только для этого поставщика
        if provider_id:
            provider = Provider.objects.filter(id=provider_id).first()
            if not provider:
                return Response({'error': 'Поставщик с указанным ID не найден.'}, status=status.HTTP_404_NOT_FOUND)

            # Получаем все записи из таблицы Invoice для данного поставщика в заданном промежутке дат
            invoices_within_range = Invoice.objects.filter(providers=provider,
                                                           product_date__range=(start_date, end_date))

            expenses_within_range = Expense.objects.filter(provider=provider,
                                                           expense_date__range=(start_date, end_date))

            # Получаем список номеров счетов
            invoice_number = list(invoices_within_range.values_list('invoice_number', flat=True))

            product_date = list(invoices_within_range.values_list('product_date', flat=True))

            product_price = list(invoices_within_range.values_list('product_price', flat=True))

            attribute = list(invoices_within_range.values_list('attribute', flat=True))

            expense_date = list(expenses_within_range.values_list('expense_date', flat=True))

            expense_number = list(expenses_within_range.values_list('expense_number', flat=True))

            expense_sum = list(expenses_within_range.values_list('expense_sum', flat=True))

            # Сумма доходов для указанного поставщика за указанный период
            invoice_total = invoices_within_range.aggregate(total_invoice=Sum('product_price'))['total_invoice'] or 0

            expense_total = expenses_within_range.aggregate(total_expense=Sum('expense_sum'))['total_expense'] or 0


            balance = invoice_total - expense_total

            total_invoice_before_start_date = \
            Invoice.objects.filter(providers=provider, product_date__lt=start_date).aggregate(
                total_invoice_before_start_date=Sum('product_price'))['total_invoice_before_start_date'] or 0

            total_expense_before_start_date = \
            Expense.objects.filter(provider=provider, expense_date__lt=start_date).aggregate(
                total_expense_before_start_date=Sum('expense_sum'))['total_expense_before_start_date'] or 0



            response_data = {
                'provider_id': provider.id,
                'provider_name': provider.provider_name,
                'start_date': start_date_str,
                'end_date': end_date_str,
                'total_expense': expense_total,
                'total_expense_before_start_date': total_expense_before_start_date,
                'total_invoice': invoice_total,
                'total_invoice_before_start_date': total_invoice_before_start_date,
                'balance': balance,
                'invoice_number': invoice_number,
                'product_date': product_date,
                'product_price': product_price,
                'attribute': attribute
            }

            # Генерируем и сохраняем отчет в Excel
            excel_file = self.generate_excel_report([response_data], provider.provider_name, end_date_str, invoice_number, product_date, product_price, expense_date, expense_number, expense_sum, attribute)

            # Возвращаем файл Excel как HTTP-ответ
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report.xlsx'
            response.write(excel_file.getvalue())

            return response

        else:
            # Сумма расходов для всех поставщиков за указанный период
            expense_totals = Expense.objects.filter(expense_date__range=(start_date, end_date)).values(
                'provider').annotate(total_expense=Sum('expense_sum'))

            # Сумма доходов для всех поставщиков за указанный период
            invoice_totals = Invoice.objects.filter(product_date__range=(start_date, end_date)).values(
                'providers').annotate(total_invoice=Sum('product_price'))

            # Получаем всех поставщиков
            providers = Provider.objects.all()

            # Создаем данные для ответа
            response_data = []

            for provider in providers:
                provider_id = provider.id
                expense_total = next(
                    (item['total_expense'] for item in expense_totals if item['provider'] == provider_id), None)
                invoice_total = next(
                    (item['total_invoice'] for item in invoice_totals if item['providers'] == provider_id), None)

                if expense_total is not None or invoice_total is not None:
                    balance = (invoice_total or 0) - (expense_total or 0)
                    response_data.append({
                        'provider_id': provider_id,
                        'provider_name': provider.provider_name,
                        'start_date': start_date_str,
                        'end_date': end_date_str,
                        'total_expense': expense_total,
                        'invoice': invoice_total,
                        'balance': balance
                    })
                else:
                    response_data.append({
                        'provider_id': provider_id,
                        'provider_name': provider.provider_name,
                        'start_date': start_date_str,
                        'end_date': end_date_str,
                        'total_expense': None,
                        'invoice': None,
                        'balance': None
                    })

            # Ищем записи, которые не попадают в указанный промежуток дат
            expenses_out_of_range = Expense.objects.exclude(expense_date__range=(start_date, end_date)).filter(
                expense_date__isnull=False)
            invoices_out_of_range = Invoice.objects.exclude(product_date__range=(start_date, end_date)).filter(
                product_date__isnull=False)

            for expense in expenses_out_of_range:
                response_data.append({
                    'provider_id': expense.provider.id,
                    'provider_name': expense.provider.provider_name,
                    'expense_date': expense.expense_date.strftime('%d.%m.%Y'),
                    'expense_sum': expense.expense_sum,
                    'product_price': None,
                    'balance': None
                })

            for invoice in invoices_out_of_range:
                response_data.append({
                    'provider_id': invoice.providers.id,
                    'provider_name': invoice.providers.provider_name,
                    'expense_date': None,
                    'expense_sum': None,
                    'product_price': invoice.product_price,
                    'balance': None
                })

            # Генерируем и сохраняем отчет в Excel
            excel_file = self.generate_excel_report(response_data, '', end_date_str, [])

            # Возвращаем файл Excel как HTTP-ответ
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=REPORT.xlsx'
            response.write(excel_file.getvalue())

            return response

    def generate_excel_report(self, data, provider_name, end_date_str, invoice_number, product_date, product_price,
                              expense_date, expense_number, expense_sum, attribute):
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Создаем ячейки для каждой строки в шапке и объединяем их
        ws.merge_cells('A1:F5')
        header_text = f'АКТ\nсверки взаимных расчетов\nмежду  MB "Alesta LT"\nи  {provider_name}\nпо состоянию на {end_date_str}'
        ws['A1'] = header_text

        # Устанавливаем выравнивание по центру
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Устанавливаем выравнивание по центру для заголовков столбцов
        for row in ws.iter_rows(min_row=1, max_row=4, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Заголовки столбцов
        ws['A6'] = provider_name
        ws.merge_cells('A6:C6')
        ws['A6'].font = Font(bold=True)

        ws['D6'] = 'MB ALESTA LT'
        ws.merge_cells('D6:F6')
        ws['D6'].font = Font(bold=True)

        ws['A7'] = 'Дата'
        ws['B7'] = 'Документ'
        ws.column_dimensions['B'].width = 15
        ws['C7'] = 'Сумма'
        ws['D7'] = 'Дата'
        ws['E7'] = 'Документ'
        ws.column_dimensions['E'].width = 15
        ws['F7'] = 'Сумма'


        # Устанавливаем выравнивание для заголовков столбцов
        for col in range(1, 7):
            header_cell = ws.cell(row=7, column=col)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            if col in [2, 5]:
                ws.row_dimensions[6].height = 30

        # Заполнение таблицы данными
        row_num = 8
        for item in data:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            ws.cell(row=row_num, column=1, value='Сальдо на ' + item.get('start_date'))
            ws.cell(row=row_num, column=3, value=item.get('total_invoice_before_start_date'))
            ws.column_dimensions['A'].width = 15

            ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=5)
            ws.cell(row=row_num, column=4, value='Сальдо на ' + item.get('start_date'))
            ws.cell(row=row_num, column=6, value=item.get('total_expense_before_start_date'))
            ws.column_dimensions['D'].width = 15

            # Увеличиваем row_num только после добавления одной записи
            row_num += 1

        # Определяем максимальное количество записей среди всех типов данных
        max_records = max(len(invoice_number), len(product_date), len(expense_date), len(attribute))

        for i in range(max_records):
            # Добавляем записи для счетов
            if i < len(invoice_number):
                ws.cell(row=row_num, column=2, value=invoice_number[i])
            # Добавляем записи для продуктовых данных
            if i < len(product_date):
                ws.cell(row=row_num, column=1, value=product_date[i])
                ws.cell(row=row_num, column=3, value=product_price[i])
            # Добавляем записи для расходов
            if i < len(expense_date):
                ws.cell(row=row_num, column=4, value=expense_date[i])
                ws.cell(row=row_num, column=5, value=expense_number[i])
                ws.cell(row=row_num, column=6, value=expense_sum[i])
            # Добавляем атрибуты
            if i < len(attribute):
                if attribute[i] == 'payment':
                    ws.cell(row=row_num, column=2, value='Оплата')
                elif attribute[i] == 'cash':
                    ws.cell(row=row_num, column=2, value=invoice_number[i])
            row_num += 1

        # Добавляем строку с итогами
        row_num += 1
        ws.cell(row=row_num, column=1, value='ИТОГО')
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        ws.cell(row=row_num, column=3, value=sum(item.get('total_invoice') or 0 for item in data))
        ws.cell(row=row_num, column=4, value='ИТОГО')
        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=5)
        ws.cell(row=row_num, column=6, value=sum(item.get('total_expense') or 0 for item in data))
        ws.merge_cells(f'A{row_num + 2}:F{row_num + 2}')
        # Вычисляем сумму всех счетов
        total_expense_sum = sum(
            (item.get('total_expense') or 0) + (item.get('total_expense_before_start_date') or 0) for item in data)
        total_invoice_sum = sum(
            (item.get('total_invoice') or 0) + (item.get('total_invoice_before_start_date') or 0) for item in data)

        difference = total_expense_sum - total_invoice_sum
        ws[
            f'A{row_num + 2}'] = f'Сальдо на {end_date_str} составляет {difference:.2f} EUR в пользу {"MB ALESTA" if sum(item.get("balance") or 0 for item in data) >= 0 else provider_name}'

        ws.merge_cells(f'A{row_num + 3}:E{row_num + 3}')
        ws[f'A{row_num + 3}'] = f'ИТОГО'

        total_cell = ws[f'A{row_num + 3}']
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, 6):
            ws.cell(row=row_num + 3, column=col).font = Font(bold=True)

        border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                        bottom=Side(style='thick'))
        for row in ws.iter_rows(min_row=row_num + 3, max_row=row_num + 3, min_col=1, max_col=5):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=1, max_row=row_num + 3, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))


        #Добавляем баланс параллеьно ИТОГО
        total_balance = difference

        ws.cell(row=row_num + 3, column=6, value=total_balance)

        total_balance_cell = ws.cell(row=row_num + 3, column=6)
        total_balance_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_balance_cell.font = Font(bold=True)

        ws.merge_cells(f'A{row_num + 4}:F{row_num + 6}')

        # Добавялем пробелы между словами
        extra_spaces = " " * 50
        extra_spaces_1 = " " * 70
        extra_spaces_2 = " " * 25

        header_text = f'{provider_name}{extra_spaces}MB "Alesta LT"'

        header_text += f'\n{extra_spaces_1} {"Директор":20}{extra_spaces_2}{"Robert Zinis"}'

        ws[f'A{row_num + 4}'] = header_text

        alignment = Alignment(vertical='center', wrap_text=True)

        ws[f'A{row_num + 4}'].alignment = alignment

        border = Border(bottom=Side(style='thick'))

        for row in ws.iter_rows(min_row=row_num + 6, max_row=row_num + 6, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        output = BytesIO()
        wb.save(output)

        output.seek(0)

        return output

class ProviderViewSet(viewsets.ModelViewSet):
    queryset = Provider.objects.all().order_by('-id')
    serializer_class = ProviderSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['provider_name'] #?search
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):          #сериализатор для POST и PATCH-запросов (GET показывает весь список, POST оставляет ID ключа другой таблицы)
        if self.request.method == 'POST':
            return ProviderCreateSerializer
        elif self.request.method == 'PATCH':
            return ProviderCreateSerializer
        return ProviderSerializer



    # Еслли в накладой есть поставщик - TRUE, иначе FALSE
    # /api/providers/5/check_invoice
    @action(detail=True, methods=['GET'])
    def check_invoice(self, request, pk=None):
        providers = self.get_object()
        invoice_exists = providers.invoice_set.exists()
        return Response({'exists': invoice_exists})

    # Поиск поставшика по букве в независимости от регистра и сортировка 5 по ID
    # api/providers/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        #Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            providers = Provider.objects.all().order_by('-id')  # Возвращает все записи, если 'name' не задан
        else:
            providers = Provider.objects.filter(provider_name__iregex=query).order_by('-id')

        # Применяем пагинацию к результатам
        paginated_providers = paginator.paginate_queryset(providers, request)

        serializer = ProviderSerializer(paginated_providers, many=True)
        return paginator.get_paginated_response(serializer.data)

    # Сортировка sort= provider_name / bank_name
    # /api/providers/filter_by_sort/?provider_name=true (false)  /  bank_name
    @action(detail=False, methods=['GET'])
    def filter_by_sort(self, request, *args, **kwargs):
        provider_name_order = self.request.query_params.get('provider_name', None)
        bank_name_order = self.request.query_params.get('bank_name', None)


        queryset = self.get_queryset()

        if provider_name_order is not None:
            # Сортировка по имени поставщика (от А до Я или от Я до А)
            provider_name_ordering = '' if provider_name_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{provider_name_ordering}provider_name')

        if bank_name_order is not None:
            # Сортировка по дате (от ранней к поздней или от поздней к ранней)
            bank_name_ordering = '' if bank_name_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{bank_name_ordering}bank__bank_name')

        paginator = PageNumberPagination()
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all().order_by('-id')
    serializer_class = InvoiceSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['invoice_number', 'provider__provider_name']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):          #сериализатор для POST и PATCH-запросов (GET показывает весь список, POST оставляет ID ключа другой таблицы)
        if self.request.method == 'POST':
            return InvoiceCreateSerializer
        elif self.request.method == 'PATCH':
            return InvoiceCreateSerializer
        return InvoiceSerializer



    # Запрос на поиск записей по атрибуту и сортировка по дате (сразу выводим записи с атрибутом и потом все остальные)
    # http://127.0.0.1:8000/api/invoices/filter_by_attribute/?attribute=плюс
    @action(detail=False, methods=['get'])
    def filter_by_attribute(self, request):
        attribute_value = request.query_params.get('attribute', None)
        paginator = PageNumberPagination()
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if attribute_value:
            # Фильтрация записей по атрибуту и сортировка по product_date
            attribute_invoices = Invoice.objects.filter(attribute=attribute_value).order_by('product_date')
            serialized_attribute_invoices = InvoiceSerializer(attribute_invoices, many=True).data

            # Остальные записи, отсортированные по product_date
            other_invoices = Invoice.objects.exclude(attribute=attribute_value).order_by('product_date')
            paginated_other_invoices = paginator.paginate_queryset(other_invoices, request)
            serialized_other_invoices = InvoiceSerializer(paginated_other_invoices, many=True).data

            # Возвращаем записи по атрибуту и остальные записи
            return paginator.get_paginated_response(
                {'attribute_invoices': serialized_attribute_invoices, 'other_invoices': serialized_other_invoices})
        else:
            # Если атрибут не задан, возвращаем все записи, отсортированные по product_date
            all_invoices = Invoice.objects.all().order_by('product_date')
            paginated_all_invoices = paginator.paginate_queryset(all_invoices, request)
            serialized_all_invoices = InvoiceSerializer(paginated_all_invoices, many=True).data
            return paginator.get_paginated_response({'all_invoices': serialized_all_invoices})


    # Сортировка sort= product_price / provider_name / product_date
    # /api/invoices/filter_by_sort/?product_date=true (false)  /  product_price / provider_name
    @action(detail=False, methods=['GET'])
    def filter_by_sort(self, request, *args, **kwargs):
        provider_name_order = self.request.query_params.get('provider_name', None)
        product_date_order = self.request.query_params.get('product_date', None)
        product_price_order = self.request.query_params.get('product_price', None)

        queryset = self.get_queryset()

        if provider_name_order is not None:
            # Сортировка по имени поставщика (от А до Я или от Я до А)
            provider_name_ordering = '' if provider_name_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{provider_name_ordering}providers__provider_name')

        if product_date_order is not None:
            # Сортировка по дате (от ранней к поздней или от поздней к ранней)
            product_date_ordering = '' if product_date_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{product_date_ordering}product_date', '-id')

        if product_price_order is not None:
            product_price_ordering = '' if product_price_order.lower() == 'true' else '-'
            # Преобразование CharField в IntegerField для корректной числовой сортировки
            queryset = queryset.annotate(
                product_price_int=Cast(F('product_price'), IntegerField())
            ).order_by(f'{product_price_ordering}product_price_int')

        paginator = PageNumberPagination()
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)




    # Поиск с даты по дату по двум параметрам + пагинация
    # http://127.0.0.1:8000/api/invoices/?date=false (true)
    def get_queryset(self):
        queryset = super().get_queryset()

        # Получаем параметр order_by из запроса
        order_by = self.request.query_params.get('date', None)

        # Сортируем записи в зависимости от параметра order_by
        if order_by == 'false':
            # Если указан date=false, сортируем по возрастанию даты, затем по возрастанию ID
            queryset = queryset.order_by('product_date', 'id')
        elif order_by == 'true':
            # Если указан date=true, сортируем по убыванию даты, затем по возрастанию ID
            queryset = queryset.order_by('-product_date', '-id')

        return queryset



    # Поиск в инвосе по provider_name
    # /invoices/search_by_provider_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_provider_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            invoices = Invoice.objects.all().order_by('-id')  # Возвращает все записи, если 'name' не задан
        else:
            invoices = Invoice.objects.filter(providers__provider_name__iregex=query).order_by('-id')

        # Применяем пагинацию к результатам
        paginated_invoices = paginator.paginate_queryset(invoices, request)

        serializer = InvoiceSerializer(paginated_invoices, many=True)
        return paginator.get_paginated_response(serializer.data)





    # Поиск invoice по invoice_number в независимости от регистра и сортировка 5 по ID
    # api/invoices/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            invoices = Invoice.objects.all().order_by('-id')  # Возвращает все записи, если 'name' не задан
        else:
            invoices = Invoice.objects.filter(invoice_number__iregex=query).order_by('-id')

        # Применяем пагинацию к результатам
        paginated_invoices = paginator.paginate_queryset(invoices, request)

        serializer = InvoiceSerializer(paginated_invoices, many=True)
        return paginator.get_paginated_response(serializer.data)



    # Поиск по параметрам : 1) даты с по, 2) дата с по + поставщик
    # api/invoices/filter_by_invoice/?start_date=2023-01-01&end_date=2023-12-31&provider_id=10
    @action(detail=False, methods=['get'])
    def filter_by_invoice(self, request):
        queryset = self.get_queryset()

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        provider_id = request.query_params.get('provider_id')

        if start_date:
            queryset = queryset.filter(product_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(product_date__lte=end_date)

        if provider_id:
            queryset = queryset.filter(providers_id__id=provider_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)



     # Поиск с даты по дату и сумму по каждому поставщику
     # http://127.0.0.1:8000/api/invoices/search_by_date_sum/?start_date=01.01.2024&end_date=31.12.2024
    @action(detail=False, methods=['GET'])
    def search_by_date_sum(self, request):
        # Получаем значения параметров 'start_date' и 'end_date' из запроса
        start_date_str = self.request.query_params.get('start_date', None)
        end_date_str = self.request.query_params.get('end_date', None)

        # Если оба параметра отсутствуют, возвращаем все строки без фильтрации по дате
        if not start_date_str and not end_date_str:
            result = (
                Provider.objects
                .annotate(
                    total_price=Sum('invoice__product_price'),
                    last_product_date=Max('invoice__product_date')
                )
                .filter(
                    Q(total_price__gt=0) |
                    Q(last_product_date__isnull=False)
                )
                .values('id', 'provider_name', 'total_price', 'last_product_date')
            )

            # Возвращаем ответ с данными
            response_data = []
            for entry in result:
                provider_id = entry['id']
                invoices = Invoice.objects.filter(providers_id=provider_id)
                total_product_price = sum(
                    Decimal(invoice.product_price) for invoice in invoices
                )
                response_data.append({
                    'provider_id': provider_id,
                    'provider_name': entry['provider_name'],
                    'product_price': '{:.2f}'.format(total_product_price),
                    'start_date': None,
                    'end_date': None,
                })

            return Response(response_data)

        # Проверяем, что оба параметра 'start_date' и 'end_date' предоставлены
        if not start_date_str or not end_date_str:
            return Response({'error': 'Необходимо указать оба параметра start_date и end_date.'}, status=400)

        try:
            # Преобразовываем значения параметров в даты
            start_date = timezone.datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = timezone.datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            # В случае некорректного формата даты возвращаем ошибку
            return Response({'error': 'Неверный формат даты. Пожалуйста, используйте DD.MM.YYYY.'}, status=400)

        # Выполняем запрос к базе данных для получения данных о поставщиках
        result = (
            Provider.objects
            .annotate(
                total_price=Sum(
                    'invoice__product_price',
                    filter=Q(invoice__product_date__range=(start_date, end_date))
                ),
                last_product_date=Max('invoice__product_date')
            )
            .filter(
                Q(total_price__gt=0) |
                Q(last_product_date__range=(start_date, end_date))
            )
            .values('id', 'provider_name', 'total_price', 'last_product_date')
        )

        # Подготавливаем данные для ответа
        response_data = []

        # Обрабатываем результаты запроса
        for entry in result:
            provider_id = entry['id']

            # Получаем счета для текущего поставщика в указанном диапазоне дат
            invoices = Invoice.objects.filter(providers_id=provider_id)

            # Преобразование CharField в числа и суммирование для продуктов в диапазоне введенной даты
            total_product_price = sum(
                Decimal(invoice.product_price) if (
                        invoice.product_date and start_date <= invoice.product_date <= end_date
                ) else Decimal(0)
                for invoice in invoices
            )

            # Добавляем данные в список ответа
            response_data.append({
                'provider_id': provider_id,
                'provider_name': entry['provider_name'],
                'product_price': '{:.2f}'.format(total_product_price),
                'start_date': start_date.strftime('%d.%m.%Y') if start_date else None,
                'end_date': end_date.strftime('%d.%m.%Y') if end_date else None,
            })

        return Response(response_data)

    # Запрос соединения выше до даты выводим сумму по данному поставщику. за промежуток выводим строки
    # http://127.0.0.1:8000/api/invoices/search_by_criteria/?start_date=10.01.2024&end_date=15.01.2024&provider_id=2
    @action(detail=False, methods=['GET'])
    def search_by_criteria(self, request):
        start_date_str = self.request.query_params.get('start_date', None)
        end_date_str = self.request.query_params.get('end_date', None)
        provider_id = self.request.query_params.get('provider_id', None)

        if not start_date_str or not end_date_str or not provider_id:
            return Response(
                {'error': 'Please provide valid start_date, end_date, and provider_id in the query parameters.'},
                status=400)

        try:
            start_date = timezone.datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = timezone.datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Please use DD.MM.YYYY.'}, status=400)

        # Получаем сумму для каждого поставщика до start_date
        total_invoice_before_start_date = (
            Provider.objects
            .filter(id=provider_id)
            .annotate(
                total_invoice=Sum('invoice__product_price', filter=Q(invoice__product_date__lt=start_date))
            )
            .values('id', 'provider_name', 'total_invoice')
            .first()
        )

        # Получаем все строки для указанного поставщика в заданном диапазоне дат
        invoices = (
            Invoice.objects
            .filter(providers_id=provider_id,
                    product_date__range=(start_date, end_date))
            .order_by('product_date')
            .values('product_date', 'product_price', 'invoice_number')
        )

        response_data = []

        # Если есть сумма до start_date, добавляем ее в ответ
        if total_invoice_before_start_date:
            response_data.append({
                'provider_id': provider_id,
                'provider_name': total_invoice_before_start_date['provider_name'],
                'total_invoice_before_start_date': total_invoice_before_start_date['total_invoice'],
                'start_date': start_date_str
            })

        # Добавляем строки с expense_sum и expense_number в промежутке между start_date и end_date
        for invoice in invoices:
            response_data.append({
                'provider_id': provider_id,
                'provider_name': total_invoice_before_start_date['provider_name'],
                'product_date': invoice['product_date'],
                'product_price': Decimal(invoice['product_price']) if invoice['product_price'] else None,
                'invoice_number': invoice['invoice_number']
            })

        return Response(response_data)






class ClientViewSet(viewsets.ModelViewSet):  #справочник
    queryset = Client.objects.all().order_by('-id')
    serializer_class = ClientSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['client_name']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]


    def get_serializer_class(self):
        #сериализатор для POST и PATCH-запросов (GET показывает весь список, POST оставляет ID ключа другой таблицы)
        if self.request.method == 'POST':
            return ClientCreateSerializer
        elif self.request.method == 'PATCH':
            return ClientCreateSerializer
        return ClientSerializer

    # Поиск клиента по букве в независимости от регистра и сортировка 5 по ID
    # api/clients/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            clients = Client.objects.all().order_by('-id')  # Возвращает все записи, если 'name' не задан
        else:
            clients = Client.objects.filter(client_name__iregex=query).order_by('-id')

        # Применяем пагинацию к результатам
        paginated_clients = paginator.paginate_queryset(clients, request)

        serializer = ClientSerializer(paginated_clients, many=True)
        return paginator.get_paginated_response(serializer.data)


class ProductViewSet(viewsets.ModelViewSet):  #справочник
    queryset = Product.objects.all().order_by('-id')
    serializer_class = ProductSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['product_name']  #?search=AAAA
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]


    # Поиск товара по букве в независимости от регистра и сортировка 5 по ID
    # api/products/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()
        paginator.page_size = 10  # Установите желаемый размер страницы

        if query:
            # Если запрос не пустой, разбиваем его на слова
            words = query.split()

            # Формируем Q объекты для каждого слова
            queries = [Q(product_name__iregex=word) for word in words]

            # Используем functools.reduce для объединения Q объектов с оператором 'и'
            combined_query = functools.reduce(lambda x, y: x & y, queries)

            # Фильтруем товары по объединенному запросу
            products = Product.objects.filter(combined_query).order_by('-id')
        else:
            # Если запрос пустой, выводим все записи
            products = Product.objects.all().order_by('-id')

        # Применяем пагинацию к результатам
        paginated_products = paginator.paginate_queryset(products, request)

        serializer = ProductSerializer(paginated_products, many=True)

        # Вручную создаем ответ, включая информацию о пагинации
        response_data = {
            'count': paginator.page.paginator.count,
            'next': paginator.get_next_link(),
            'previous': paginator.get_previous_link(),
            'results': serializer.data
        }

        return Response(response_data, status=status.HTTP_200_OK)


class GroupViewSet(viewsets.ModelViewSet): #справочник
    queryset = Group.objects.all().order_by('-id')
    serializer_class = GroupSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['group_name']   #?search=AAAA
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]


class IncomeViewSet(viewsets.ModelViewSet):
    serializer_class = IncomeSerializer
    queryset = Income.objects.all().order_by('-id')
    filter_backends = [filters.SearchFilter]
    search_fields = [] #?search=AAAA
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]



    def get_serializer_class(
            self):  # сериализатор для POST и PACT-запросов (GET показывает весь список, POST оставляет ID ключа другой таблицы)
        if self.request.method == 'POST':
            return IncomeCreateSerializer
        elif self.request.method == 'PATCH':
            return IncomeCreateSerializer
        return IncomeSerializer



    # Поиск с даты по дату по двум параметрам + пагинация
    # http://127.0.0.1:8000/api/incomes/?date=min (max)
    def get_queryset(self):
        queryset = super().get_queryset()

        # Получаем параметр order_by из запроса
        order_by = self.request.query_params.get('date', None)

        # Сортируем записи в зависимости от параметра order_by
        if order_by == 'min':
            # Если указан date=min, сортируем по возрастанию
            queryset = queryset.order_by('invoice__product_date')
        elif order_by == 'max':
            # Если указан date=max, сортируем по убыванию
            queryset = queryset.order_by('-invoice__product_date')

        return queryset




     # Поиск получить где параметрами будут : 1) даты с по, 2) дата с по + поставщик, 3) дата с по + товар, 4)дата с по + поставщик + товар
     # incomes/filter_by_income/?start_date=2023-01-01&end_date=2023-12-31&providers_id=2&product_id=1
    @action(detail=False, methods=['get'])
    def filter_by_income(self, request):
        queryset = self.get_queryset()

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        provider_id = request.query_params.get('providers_id')
        product_id = request.query_params.get('product_id')

        filters = Q()

        if start_date and end_date:
            filters &= Q(invoice__product_date__range=(start_date, end_date))
        elif start_date:
            filters &= Q(invoice__product_date__gte=start_date)
        elif end_date:
            filters &= Q(invoice__product_date__lte=end_date)

        if provider_id:
            filters &= Q(invoice__providers_id=provider_id)

        if product_id:
            filters &= Q(product_id=product_id)

        queryset = queryset.filter(filters)

        serializer = IncomeSerializer(queryset, many=True)
        return Response(serializer.data)




    #Поиск income по invoice_id от даты и по дату
    # api/incomes/filter_by_date_and_invoice/?invoice_id=476&start_date=2023-11-01&end_date=2023-12-31
    @action(detail=False, methods=['GET'])
    def filter_by_date_and_invoice(self, request):
        queryset = self.get_queryset()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))

        invoice_id = request.query_params.get('invoice_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if invoice_id:
            queryset = queryset.filter(invoice__id=invoice_id)

        if start_date:
            queryset = queryset.filter(invoice__product_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(invoice__product_date__lte=end_date)

        # Делаем пагинацию
        paginator = PageNumberPagination()
        paginator.page_size = page_size
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = self.get_serializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)




    # Поиск income по product_name от даты и по дату
    # api/incomes/filter_by_date_and_product_name/?product_name=476&start_date=2023-11-01&end_date=2023-12-31
    @action(detail=False, methods=['GET'])
    def filter_by_date_and_product_name(self, request):
        queryset = self.get_queryset()

        product_name = request.query_params.get('product_name')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if product_name:
            queryset = queryset.filter(product__product_name__iregex=product_name)

        if start_date:
            queryset = queryset.filter(invoice__product_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(invoice__product_date__lte=end_date)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class BankViewSet(viewsets.ModelViewSet): #справочник
    queryset = Bank.objects.all().order_by('-id')
    serializer_class = BankSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['bank_name']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]



    # Поиск банка по букве в независимости от регистра и сортировка 5 по ID
    # api/banks/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            banks = Bank.objects.all().order_by('-id')  # Возвращает все записи, если 'name' не задан
        else:
            banks = Bank.objects.filter(bank_name__iregex=query).order_by('-id')

        # Применяем пагинацию к результатам
        paginated_banks = paginator.paginate_queryset(banks, request)

        serializer = BankSerializer(paginated_banks, many=True)
        return paginator.get_paginated_response(serializer.data)


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all().order_by('-id')
    serializer_class = ExpenseSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(
            self):  # сериализатор для POST и PATCH-запросов (GET показывает весь список, POST оставляет ID ключа другой таблицы)
        if self.request.method == 'POST':
            return ExpenseCreateSerializer
        elif self.request.method == 'PATCH':
            return ExpenseCreateSerializer
        return ExpenseSerializer

    # Поиск с даты по дату по двум параметрам + пагинация
    # http://127.0.0.1:8000/api/expenses/?date=false (true)
    def get_queryset(self):
        queryset = super().get_queryset()

        # Получаем параметр order_by из запроса
        order_by = self.request.query_params.get('date', None)

        # Сортируем записи в зависимости от параметра order_by
        if order_by == 'false':
            # Если указан date=false, сортируем по возрастанию даты, затем по возрастанию ID
            queryset = queryset.order_by('expense_date', 'id')
        elif order_by == 'true':
            # Если указан date=true, сортируем по убыванию даты, затем по убыванию ID
            queryset = queryset.order_by('-expense_date', '-id')

        return queryset


    # Сортировка sort= product_price / provider_name / product_date
    # /api/expenses/filter_by_sort/?expense_date=True (False)  /  expense_sum / provider_name
    @action(detail=False, methods=['GET'])
    def filter_by_sort(self, request, *args, **kwargs):
        # Получение параметров запроса для сортировки
        provider_name_order = self.request.query_params.get('provider_name', None)
        expense_date_order = self.request.query_params.get('expense_date', None)
        expense_sum_order = self.request.query_params.get('expense_sum', None)

        # Получение изначального набора данных
        queryset = self.get_queryset().order_by('-id')

        # Если указан параметр provider_name, выполняется сортировка по имени поставщика
        if provider_name_order is not None:
            provider_name_ordering = '' if provider_name_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{provider_name_ordering}provider__provider_name')

        # Если указан параметр expense_date, выполняется сортировка по дате
        if expense_date_order is not None:
            expense_date_ordering = '' if expense_date_order.lower() == 'true' else '-'
            queryset = queryset.order_by(f'{expense_date_ordering}expense_date', '-id')

        # Если указан параметр expense_sum, выполняется сортировка по сумме расхода
        if expense_sum_order is not None:
            expense_sum_ordering = '' if expense_sum_order.lower() == 'true' else '-'
            # Преобразование CharField в IntegerField для корректной числовой сортировки
            queryset = queryset.annotate(
                expense_sum_int=Cast(F('expense_sum'), IntegerField())
            ).order_by(f'{expense_sum_ordering}expense_sum_int')

        # Настройка пагинации
        paginator = PageNumberPagination()
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        # Получение отфильтрованного и пагинированного набора данных
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        # Сериализация данных и возврат ответа с пагинацией
        serializer = self.get_serializer(paginated_queryset, many=True)
        return paginator.get_paginated_response(serializer.data)


    # Поиск expense по expense_number и provider_name букве в независимости от регистра и сортировка 5 по ID
    # api/expenses/filter_by_keyword/?query=234+вадим и наоборот

    @action(detail=False, methods=['GET'])
    def filter_by_keyword(self, request):
        query = request.query_params.get('query', '')
        terms = query.split(' ')  # Разбиваем запрос на отдельные термины
        queryset = Expense.objects.all().order_by('-id')

        q_objects = Q()
        for term in terms:
            q_objects &= (
                    Q(expense_number__icontains=term) |  # Поиск по expense_number
                    Q(provider__provider_name__iregex=term)  # Поиск по provider_name
            )

        results = queryset.filter(q_objects)

        self.pagination_class = APIListPagination
        page = self.paginate_queryset(results)

        if page is not None:
            serializer = ExpenseSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ExpenseSerializer(results, many=True)
        return Response({'results': serializer.data})



    # Поиск по параметрам : 1) даты с по, 2) дата с по + поставщик
    # expenses/filter_by_expense/?start_date=2023-01-01&end_date=2023-12-31&provider_id=10
    @action(detail=False, methods=['get'])
    def filter_by_expense(self, request):
        queryset = self.get_queryset()

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        provider_id = request.query_params.get('provider_id')


        if start_date:
            queryset = queryset.filter(expense_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(expense_date__lte=end_date)

        if provider_id:
            queryset = queryset.filter(provider__id=provider_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)



      #поиск по поставщику и сумма expense_sum

    # Запрос про сумму по каждому поставщику с даты по дату, если дата пустая выводит всех поставщиков
    # http://127.0.0.1:8000/api/expenses/search_by_date_sum/?start_date=01.01.2024&end_date=31.12.2024
    @action(detail=False, methods=['GET'])
    def search_by_date_sum(self, request):
        # Получаем значения параметров 'start_date' и 'end_date' из запроса
        start_date_str = self.request.query_params.get('start_date', None)
        end_date_str = self.request.query_params.get('end_date', None)

        # Если оба параметра отсутствуют, возвращаем все строки без фильтрации по дате
        if not start_date_str and not end_date_str:
            result = (
                Provider.objects
                .annotate(
                    total_price=Sum('expense__expense_sum'),
                    last_expense_date=Max('expense__expense_date')
                )
                .filter(
                    Q(total_price__gt=0) |
                    Q(last_expense_date__isnull=False)
                )
                .values('id', 'provider_name', 'total_price', 'last_expense_date')
            )

            # Возвращаем ответ с данными
            response_data = []
            for entry in result:
                provider_id = entry['id']
                expenses = Expense.objects.filter(provider_id=provider_id)
                total_expense_sum = sum(
                    Decimal(expense.expense_sum) for expense in expenses
                )
                response_data.append({
                    'provider_id': provider_id,
                    'provider_name': entry['provider_name'],
                    'expense_sum': '{:.2f}'.format(total_expense_sum),
                    'start_date': None,
                    'end_date': None,
                })

            return Response(response_data)

        # Проверяем, что оба параметра 'start_date' и 'end_date' предоставлены
        if not start_date_str or not end_date_str:
            return Response({'error': 'Необходимо указать оба параметра start_date и end_date.'}, status=400)

        try:
            # Преобразовываем значения параметров в даты
            start_date = timezone.datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = timezone.datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            # В случае некорректного формата даты возвращаем ошибку
            return Response({'error': 'Неверный формат даты. Пожалуйста, используйте DD.MM.YYYY.'}, status=400)

        # Выполняем запрос к базе данных для получения данных о поставщиках
        result = (
            Provider.objects
            .annotate(
                total_price=Sum(
                    'expense__expense_sum',
                    filter=Q(expense__expense_date__range=(start_date, end_date))
                ),
                last_expense_date=Max('expense__expense_date')
            )
            .filter(
                Q(total_price__gt=0) |
                Q(last_expense_date__range=(start_date, end_date))
            )
            .values('id', 'provider_name', 'total_price', 'last_expense_date')
        )

        # Подготавливаем данные для ответа
        response_data = []

        # Обрабатываем результаты запроса
        for entry in result:
            provider_id = entry['id']

            # Получаем счета для текущего поставщика в указанном диапазоне дат
            expenses = Expense.objects.filter(provider_id=provider_id)

            # Преобразование CharField в числа и суммирование для продуктов в диапазоне введенной даты
            total_expense_sum = sum(
                Decimal(expense.expense_sum) if (
                        expense.expense_date and start_date <= expense.expense_date <= end_date
                ) else Decimal(0)
                for expense in expenses
            )

            # Добавляем данные в список ответа
            response_data.append({
                'provider_id': provider_id,
                'provider_name': entry['provider_name'],
                'expense_sum': '{:.2f}'.format(total_expense_sum),
                'start_date': start_date.strftime('%d.%m.%Y') if start_date else None,
                'end_date': end_date.strftime('%d.%m.%Y') if end_date else None,
            })

        return Response(response_data)



    # запрос соединение выше до даты выводим сумму по данному поставщику. за промежуток выводим строки
    # http://127.0.0.1:8000/api/expenses/search_by_criteria/?start_date=10.01.2024&end_date=15.01.2024&provider_id=2
    @action(detail=False, methods=['GET'])
    def search_by_criteria(self, request):
        start_date_str = self.request.query_params.get('start_date', None)
        end_date_str = self.request.query_params.get('end_date', None)
        provider_id = self.request.query_params.get('provider_id', None)

        if not start_date_str or not end_date_str or not provider_id:
            return Response(
                {'error': 'Please provide valid start_date, end_date, and provider_id in the query parameters.'},
                status=400)

        try:
            start_date = timezone.datetime.strptime(start_date_str, '%d.%m.%Y').date()
            end_date = timezone.datetime.strptime(end_date_str, '%d.%m.%Y').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Please use DD.MM.YYYY.'}, status=400)

        # Получаем сумму для каждого поставщика до start_date
        total_expense_before_start_date = (
            Provider.objects
            .filter(id=provider_id)
            .annotate(
                total_expense=Sum('expense__expense_sum', filter=Q(expense__expense_date__lt=start_date))
            )
            .values('id', 'provider_name', 'total_expense')
            .first()
        )

        # Получаем все строки для указанного поставщика в заданном диапазоне дат
        expenses = (
            Expense.objects
            .filter(provider_id=provider_id,
                    expense_date__range=(start_date, end_date))
            .order_by('expense_date')
            .values('expense_date', 'expense_sum', 'expense_number')
        )

        response_data = []

        # Если есть сумма до start_date, добавляем ее в ответ
        if total_expense_before_start_date:
            response_data.append({
                'provider_id': provider_id,
                'provider_name': total_expense_before_start_date['provider_name'],
                'total_expense_before_start_date': total_expense_before_start_date['total_expense'],
                'start_date': start_date_str
            })

        # Добавляем строки с expense_sum и expense_number в промежутке между start_date и end_date
        for expense in expenses:
            response_data.append({
                'provider_id': provider_id,
                'provider_name': total_expense_before_start_date['provider_name'],
                'expense_date': expense['expense_date'],
                'expense_sum': Decimal(expense['expense_sum']) if expense['expense_sum'] else None,
                'expense_number': expense['expense_number'],
            })

        return Response(response_data)


class Expense_itemViewSet(viewsets.ModelViewSet):
    queryset = Expense_item.objects.all().order_by('-id')
    serializer_class = Expense_itemSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return Expense_itemCreateSerializer  #сериализатор для POST-запросов (GET показывает весь список, POST оставляет ID)
        return Expense_itemDetailSerializer

    # Поиск расхода по id_накладной расхода
    # api/expenses_item/?expense_id=1 без пагинации
    def list(self, request, *args, **kwargs):
        expense_id = request.query_params.get('expense_id')

        if expense_id:
            queryset = Expense_item.objects.filter(expense_id=expense_id)
        else:
            queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)



    #Поиск по параметрам : 1) даты с по, 2) дата с по + поставщик, 3) дата с по + товар, 4)дата с по + клиент + товар
    # filter_by_expense_item/?start_date=2023-01-01&end_date=2023-12-31&provider_id=10&product_id=94
    @action(detail=False, methods=['get'])
    def filter_by_expense_item(self, request):
        queryset = self.get_queryset()

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        provider_id = request.query_params.get('provider_id')
        product_id = request.query_params.get('product_id')

        if start_date:
            queryset = queryset.filter(expense__expense_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(expense__expense_date__lte=end_date)

        if provider_id:
            queryset = queryset.filter(expense__provider__id=provider_id)

        if product_id:
            queryset = queryset.filter(product__id=product_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)



    # Поиск expense_item по product_name от даты и по дату
    # expenses_item/filter_by_date_and_product_name/?product_name=476&start_date=2023-11-01&end_date=2023-12-31
    @action(detail=False, methods=['GET'])
    def filter_by_date_and_product_name(self, request):
        queryset = self.get_queryset()

        product_name = request.query_params.get('product_name')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if product_name:
            queryset = queryset.filter(product__product_name__iregex=product_name)

        if start_date:
            queryset = queryset.filter(expense__expense_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(expense__expense_date__lte=end_date)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)




class UpdateStock(APIView):
    """ Объединение записей """

    def post(self, request, prod_id):
        try:
            prod = Stock.objects.get(pk=prod_id)
        except Stock.DoesNotExist:
            raise Http404("Товар не существует")

        with transaction.atomic():
            # Находим все записи с заданными параметрами товаров
            existing_stocks = Stock.objects.filter(
                product_id=prod.product_id,
                product_price=prod.product_price,
                product_country=prod.product_country,
                product_vendor=prod.product_vendor
            ).exclude(product_barcode="").order_by('id')

            if existing_stocks:
                # Проверяем, что не существует уже объединенной записи для данной группы товаров
                existing_merged_stock = Stock.objects.filter(
                    product_id=prod.product_id,
                    product_price=prod.product_price,
                    product_country=prod.product_country,
                    product_vendor=prod.product_vendor,
                    product_barcode__isnull=False
                ).exclude(id__in=existing_stocks).first()

                if not existing_merged_stock:
                    # Создаем список для объединения всех найденных записей
                    stocks_to_merge = existing_stocks

                    earliest_barcode = min(stock.product_barcode for stock in stocks_to_merge)

                    # Вычисляем суммы для объединения
                    total_expense_full_price = sum(float(stock.expense_full_price) for stock in stocks_to_merge)
                    total_product_quantity = sum(int(stock.product_quantity) for stock in stocks_to_merge)

                    # Проверяем количество товара
                    if total_product_quantity == 0:
                        # Если количество равно 0, сохраняем значение stock
                        last_deleted_stock = stocks_to_merge.first().id

                        # Удаляем все объединенные записи, кроме последней
                        stocks_to_merge.exclude(id=last_deleted_stock).delete()

                        # Удаление из модели Income связанных записей с удаленными Stock
                        Income.objects.filter(stock=last_deleted_stock).delete()
                    else:
                        # Создаем новую запись Stock с общими значениями
                        new_stock = Stock.objects.create(
                            product_id=prod.product_id,
                            product_price=prod.product_price,
                            product_country=prod.product_country,
                            product_vendor=prod.product_vendor,
                            product_reserve=prod.product_reserve,
                            product_price_provider=prod.product_price_provider,
                            expense_allowance=prod.expense_allowance,
                            product_vat=prod.product_vat,
                            product_barcode=earliest_barcode,
                            expense_full_price=str(total_expense_full_price),
                            product_quantity=str(total_product_quantity)
                        )

                        # Обновляем связанные записи в модели Income
                        Income.objects.filter(stock__in=stocks_to_merge).update(stock_id=new_stock.id)

                        # Удаляем все объединенные записи, кроме новой
                        stocks_to_merge.exclude(id=new_stock.id).delete()

                        # Возвращаем информацию о новой записи и объединенных записях
                        merged_stocks_info = {
                            "new_stock": {"id": new_stock.id},
                            "merged_stocks": [{"id": stock.id} for stock in stocks_to_merge]
                        }
                        return Response(merged_stocks_info)

                    return Response({})


class StockViewSet(viewsets.ModelViewSet):
    queryset = Stock.objects.all().order_by('-id')
    serializer_class = StockSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['product__product_name', 'product__id']  #?search=AAAA (имя товара / ID товара)
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Http404:
            return Response([])  # Возвращаем пустой массив, если объект не найден по ID




    def get_serializer_class(self):
        if self.action == 'create':
            return StockCreateSerializer  #сериализатор для POST-запросов (GET показывает весь список, POST оставляет ID)
        return StockDetailSerializer




    #Поиск с даты по дату + товар
    # filter_by_stock/?start_date=2023-01-01&end_date=2023-12-31&product_id=84
    @action(detail=False, methods=['get'])
    def filter_by_stock(self, request):
        queryset = self.get_queryset()

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        product_id = request.query_params.get('product_id')

        if start_date:
            queryset = queryset.filter(income__invoice__product_date__gte=start_date)

        if end_date:
            queryset = queryset.filter(income__invoice__product_date__lte=end_date)

        if product_id:
            queryset = queryset.filter(product_id=product_id)


        serializer = StockSerializer(queryset, many=True)
        return Response(serializer.data)






    # Замена строк где есть одинаковые параметры ввода и обновление последневведенных + замена id обновленного stock в таблицу income
    #@receiver(pre_save, sender=Stock) - это выше в отдельной функции class UpdateStock(APIView):



    # Поиск товаров от одного до пяти слов и артикула. Список всех товаров связанных с первым словом +  список всех товаров с вторым словом и так далее
    #/api/stocks/filter_by_keyword/?query=рамка+дом+артикул+страна в числовом значении

    @action(detail=False, methods=['GET'])
    def filter_by_keyword(self, request):
        query = request.query_params.get('query', '')
        country_filter = request.query_params.get('country', '')

        queryset = Stock.objects.all().order_by('-id')

        if query or country_filter:
            keywords = query.split()
            product_name_query = Q()

            for keyword in keywords:
                product_name_query |= Q(product__product_name__iregex=keyword)

            results = queryset.filter(
                product_name_query |
                Q(product_vendor__icontains=query) |
                Q(product_country__icontains=query)
            )

            if country_filter:
                results = results.filter(product_country__icontains=country_filter)
        else:

            results = queryset

        self.pagination_class = APIListPagination
        page = self.paginate_queryset(results)

        if page is not None:
            serializer = StockSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = StockSerializer(results, many=True)
        return Response({'results': serializer.data})



    # Поиск на складе по названию товара по любой букве без учета регистра
    # api/stocks/search_by_name/?name=а
    # Поиск по названию без учета регистра - iregex, если по любой букве - icontains
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        search_query = request.query_params.get('name', '')

        if not search_query:
            # Если запрос пуст, возвращает все строки
            stock = Stock.objects.all()
        else:
            # Поле, связанное с моделью Product для поиска по названию без учета регистра
            products = Product.objects.filter(product_name__iregex=search_query)
            stock = Stock.objects.filter(product__in=products)

        stock = stock.order_by('-product')[:5] #Упорядочиваем результаты по продукту в убывающем порядке

        if stock:
            serializer = StockSerializer(stock, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({'detail': 'Товары не найдены'}, status=status.HTTP_404_NOT_FOUND)





    # Поиск на складе товара по его ID GET
    # /api/stocks/search_by_product_id/?product_id=2
    @action(detail=False, methods=['GET'])
    def search_by_product_id(self, request):
        product_id = request.query_params.get('product_id')

        if not product_id:
            return Response({"message": "Не указан ID товара"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            stock = Stock.objects.filter(
                product__id=product_id)  # filter-выводит два одинаковых ID товара, get - один товар
        except Stock.DoesNotExist:
            return Response({"response": "false"}, status=status.HTTP_200_OK)

        serializer = self.get_serializer(stock,
                                         many=True)  # filter- выводит два одинковых ID , get - один товар(удаляем many=TRUE)
        return Response(serializer.data)


    # Поиск на складе по точному количеству товара
    # api/stocks/search_by_quantity/?quantity=0
    @action(detail=False, methods=['GET'])
    def search_by_quantity(self, request):
        quantity = request.query_params.get('quantity')

        if quantity is not None:
            try:
                quantity = int(quantity)
                products_with_exact_quantity = Stock.objects.filter(product_quantity=quantity)
                serializer = StockSerializer(products_with_exact_quantity, many=True)
                return Response(serializer.data)
            except ValueError:
                return Response({'error': 'Некорректно введено количество товара. Укажите целое число'}, status=400)


class Price_changeViewSet(viewsets.ModelViewSet):
    queryset = Price_change.objects.all().order_by('-id')
    serializer_class = Price_changeSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['income__id'] #?search
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return Price_changeCreateSerializer  #сериализатор для POST-запросов (GET показывает весь список, POST оставляет ID)
        return Price_changeDetailSerializer

    # Поиск по номеру инвойса и название товара с учетом фильтров
    # api/prices_change/search_by_invoice_product/?invoice_number=1&product_name=асфа
    @action(detail=False, methods=['GET'])
    def search_by_invoice_product(self, request):
        invoice_number = request.GET.get('invoice_number')
        product_name = request.GET.get('product_name')

        if invoice_number is not None:
            # Создаем фильтр для поиска по номеру инвойса
            invoice_filter = Q(invoice_number__iexact=invoice_number)

            # Создаем фильтр для поиска по названию товара (если указано)
            product_filter = Q()
            if product_name:
                # Переносим в нижний регистр и разделяем на слова
                product_name = product_name.lower()
                product_name_parts = product_name.split()

                # Создаем фильтра для поиска по названию товара (без учета регистра и частичное совпадение)
                for part in product_name_parts:
                    product_filter |= Q(product_name__iregex=part)

            # Искать и фильтровать записи в таблицах Invoice и Product
            invoices = Invoice.objects.filter(invoice_filter)
            products = Product.objects.filter(product_filter)

            # Если найдены соответствующие записи, вернуть их сериализованные данные
            if invoices.exists() and products.exists():
                invoice_serializer = InvoiceSerializer(invoices, many=True)
                product_serializer = ProductSerializer(products, many=True)
                return Response({
                    'invoices': invoice_serializer.data,
                    'products': product_serializer.data
                })
            else:
                return Response({'message': 'Записей не найдено.'}, status=404)
        else:
            # Если номер инвойса не указан, вернуть всю таблицу PriceChange
            price_changes = Price_change.objects.all()
            price_change_serializer = Price_changeSerializer(price_changes, many=True)
            return Response({'price_changes': price_change_serializer.data})


class RetailViewSet(viewsets.ModelViewSet):
    queryset = Retail.objects.all().order_by('-id')
    serializer_class = RetailSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['']
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return RetailCreateSerializer  #сериализатор для POST-запросов (GET показывает весь список, POST оставляет ID)
        return RetailDetailSerializer



class ContractViewSet(viewsets.ModelViewSet):
    queryset = Contract.objects.all().order_by('-id')
    serializer_class = ContractSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['client__id']    #?search=1
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return ContractCreateSerializer  #сериализатор для POST-запросов (GET показывает весь список, POST оставляет ID)
        return ContractDetailSerializer

    # Поиск договоров по номеру договора
    # api/contracts/?contract_number=123
    def get_queryset(self):
        contract_number = self.request.query_params.get('contract_number')
        if contract_number:
            return Contract.objects.filter(contract_number=contract_number)
        return super().get_queryset()



    # Поиск по двум параметрам где номер контракта по цифре все данные,или есть пустой запрос то весь список выдает
    # api/contracts/search_by_client_and_number/?provider_id=4&contract_number=1234
    @action(detail=False, methods=['GET'])
    def search_by_client_and_number(self, request):
        provider_id = request.query_params.get('provider_id')
        contract_number = request.query_params.get('contract_number')

        if not provider_id:
            return Response({"message": "Не указан параметр provider_id"}, status=status.HTTP_400_BAD_REQUEST)

        contracts = Contract.objects.filter(provider_id=provider_id)

        if contract_number:
            contracts = contracts.filter(contract_number__icontains=contract_number)

        # Сортируем записи по дате создания в обратном порядке в количестве 5 записей
        contracts = contracts.order_by('-contract_number')[:5]

        serializer = self.get_serializer(contracts, many=True)
        return Response(serializer.data)





class CountryViewSet(viewsets.ModelViewSet):  #справочник
    queryset = Country.objects.all().order_by('-id')
    serializer_class = CountrySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['country_name']  #?search=AAAA
    pagination_class = APIListPagination
    permission_classes = [IsAuthenticated]

    # Поиск страны по букве в независимости от регистра и сортировка 5 по ID
    # api/countries/search_by_name/?name=
    @action(detail=False, methods=['GET'])
    def search_by_name(self, request):
        query = request.query_params.get('name', '')

        # Делаем пагинацию
        paginator = PageNumberPagination()

        # Получаем количество записей на странице из параметра запроса, по умолчанию 5
        page_size = int(request.query_params.get('page_size', 5))
        paginator.page_size = page_size

        if not query:
            countries = Country.objects.all()  # Возвращает все записи, если 'name' не задан
        else:
            countries = Country.objects.filter(country_name__iregex=query).order_by('-id') #Поиск по имени по любой букве

        # Применяем пагинацию к результатам
        paginated_countries = paginator.paginate_queryset(countries, request)

        serializer = CountrySerializer(paginated_countries, many=True)
        return paginator.get_paginated_response(serializer.data)












